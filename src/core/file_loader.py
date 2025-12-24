import pandas as pd
import openpyxl
import xlrd
import os
import traceback

# --- PART 1: PUBLIC ENTRY POINTS ---

def load_bom_file(file_path):
    print(f"\n[DEBUG] Loading BOM: {file_path}")
    return _load_generic(file_path, mode="BOM")

def load_xy_file(file_path):
    print(f"\n[DEBUG] Loading XY: {file_path}")
    return _load_generic(file_path, mode="XY")

# --- PART 2: INTERNAL LOGIC ---

def _load_generic(file_path, mode):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    ext = os.path.splitext(file_path)[1].lower()
    print(f"[DEBUG] File Extension: {ext}")
    
    # 1. Detect Start Row (Only for XY)
    header_row = 0
    if mode == "XY":
        header_row = _find_xy_header_row(file_path, ext)
        print(f"[DEBUG] Header Hunter found start row: {header_row}")
    
    try:
        # 2. Route to correct processor
        if ext in ['.xlsx', '.xlsm']:
            df = _process_xlsx(file_path, header_row, unmerge=(mode=="BOM"))
        elif ext == '.xls':
            df = _process_xls(file_path, header_row, unmerge=(mode=="BOM"))
        elif ext in ['.csv', '.txt']:
            df = _process_csv(file_path, header_row)
        else:
            raise ValueError(f"Unsupported format: {ext}")

        # 3. Cleanup
        if df is None or df.empty:
            print("[DEBUG] Warning: DataFrame is empty after loading.")
            return pd.DataFrame() # Return empty DF instead of crashing

        # Clean headers
        df.columns = df.columns.astype(str).str.strip()
        print(f"[DEBUG] Columns Found: {list(df.columns)}")
        
        # Drop completely empty rows
        original_len = len(df)
        df.dropna(how='all', inplace=True)
        print(f"[DEBUG] Rows: {original_len} -> {len(df)} (after dropping empty)")
        
        return df

    except Exception as e:
        # Print full error to console for debugging
        print("!!! LOAD ERROR !!!")
        traceback.print_exc()
        raise RuntimeError(f"Failed during load: {str(e)}")

def _find_xy_header_row(file_path, ext):
    """Scans the first 50 lines to find the true header row."""
    keywords = ["designator", "ref", "component", "layer", "side", "rotation", "mid x", "center-x", "x-loc", "part"]
    
    lines = []
    try:
        if ext in ['.csv', '.txt']:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = [f.readline() for _ in range(50)]
        else:
            # Excel Peek
            if ext == '.xls':
                wb = xlrd.open_workbook(file_path, formatting_info=False, on_demand=True)
                ws = wb.sheet_by_index(0)
                for r in range(min(50, ws.nrows)):
                    row_vals = [str(c) for c in ws.row_values(r)]
                    lines.append(" ".join(row_vals))
            else: # xlsx
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                for i, row in enumerate(ws.iter_rows(max_row=50, values_only=True)):
                    row_vals = [str(c) if c is not None else "" for c in row]
                    lines.append(" ".join(row_vals))
                wb.close()
    except Exception as e:
        print(f"[DEBUG] Header Hunter failed to peek file: {e}")
        return 0 # Fallback to 0

    # Score each line
    best_row = 0
    max_score = 0
    
    for idx, line in enumerate(lines):
        score = 0
        line_lower = str(line).lower()
        for k in keywords:
            if k in line_lower:
                score += 1
        
        if score > max_score and score >= 2:
            max_score = score
            best_row = idx
            
    return best_row

# --- PART 3: FILE PROCESSORS ---

def _process_xlsx(file_path, header_row, unmerge):
    print("[DEBUG] Processing XLSX...")
    # Add data_only=True to get values, BUT handle if it fails
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        raise ValueError(f"OpenPyXL failed to open file: {e}")

    sheet = wb.active

    if unmerge:
        print("[DEBUG] Unmerging cells...")
        merged_ranges = list(sheet.merged_cells.ranges)
        for merged_cell in merged_ranges:
            min_col, min_row, max_col, max_row = merged_cell.bounds
            top_left_value = sheet.cell(row=min_row, column=min_col).value
            sheet.unmerge_cells(str(merged_cell))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = top_left_value

    data = list(sheet.values)
    
    # Check if data exists
    if not data:
        print("[DEBUG] XLSX sheet has no data.")
        return pd.DataFrame()

    # Slice off noise
    if header_row < len(data):
        data = data[header_row:]
    else:
        print(f"[DEBUG] Header Row {header_row} is beyond file length {len(data)}. Resetting to 0.")
        header_row = 0
        
    if not data: return pd.DataFrame()

    cols = data[0]
    # Verify cols is not None
    if cols is None: cols = []
    
    # Create DF
    # Handle case where file has only 1 row (Header)
    if len(data) == 1:
        return pd.DataFrame(columns=cols)
        
    df = pd.DataFrame(data[1:], columns=cols)
    return df

def _process_xls(file_path, header_row, unmerge):
    print("[DEBUG] Processing XLS...")
    book = xlrd.open_workbook(file_path, formatting_info=unmerge)
    sheet = book.sheet_by_index(0)

    data = []
    for r in range(sheet.nrows):
        data.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])

    if unmerge:
        for crange in sheet.merged_cells:
            rlo, rhi, clo, chi = crange
            top_val = data[rlo][clo]
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    data[r][c] = top_val

    if header_row < len(data):
        data = data[header_row:]
    
    if not data: return pd.DataFrame()

    cols = data[0]
    df = pd.DataFrame(data[1:], columns=cols)
    return df

def _process_csv(file_path, header_row):
    print("[DEBUG] Processing CSV/TXT...")
    encodings = ['utf-8', 'cp1252', 'latin1']
    delimiter = ','
    if file_path.endswith('.txt'): delimiter = '\t'

    for enc in encodings:
        try:
            print(f"[DEBUG] Trying encoding: {enc}, skip: {header_row}")
            df = pd.read_csv(file_path, sep=delimiter, encoding=enc, header=0, skiprows=header_row, dtype=str)
            return df
        except Exception as e:
            # Continue to next encoding, but print warning
            pass
            
    raise ValueError("Could not decode CSV/TXT file. Unknown encoding.")